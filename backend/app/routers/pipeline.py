import json
import os
import io
import csv
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from .. import models, schemas
from ..pipeline_agents import (
    agent_parse,
    agent_screen,
    agent_deep_rank,
    agent_finalize,
)

router = APIRouter(prefix="/pipeline", tags=["pipeline"])


@router.post("/run", response_model=schemas.PipelineFullOut)
def run_pipeline(payload: schemas.PipelineRunCreate, db: Session = Depends(get_db)):
    job_title = payload.job_title or "Software Engineer"
    job_description = payload.job_description or ""
    candidate_ids = payload.candidate_ids

    if not candidate_ids:
        all_candidates = db.query(models.Candidate).order_by(models.Candidate.created_at.desc()).limit(50).all()
        candidate_ids = [c.id for c in all_candidates]

    if not candidate_ids:
        raise HTTPException(400, "No candidates found to run the pipeline. Fetch or upload candidates first.")

    run = models.PipelineRun(
        job_title=job_title,
        job_description=job_description,
        status="running",
        total_candidates=len(candidate_ids),
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    candidates_map = {}
    for cid in candidate_ids:
        c = db.query(models.Candidate).filter(models.Candidate.id == cid).first()
        if c:
            candidates_map[cid] = c

    working_data = []

    try:
        run.current_agent = "Parser Agent"
        run.progress = 10
        db.commit()

        for idx, cid in enumerate(candidate_ids):
            c = candidates_map.get(cid)
            if not c:
                continue
            cand_dict = {
                "id": c.id,
                "name": c.name,
                "email": c.email,
                "role": c.role,
                "cv_text": c.cv_text or "",
                "skills": c.skills or "",
                "experience_years": c.experience_years,
                "location": c.location or "",
                "gender": c.gender,
                "shift_preference": c.shift_preference,
                "is_remote": c.is_remote,
                "age": c.age,
                "summary": c.summary or "",
            }
            parsed = agent_parse(cand_dict, job_description)
            working_data.append({"candidate_id": cid, "candidate": cand_dict, "parsed": parsed})
            run.parsed_count = idx + 1
            run.progress = 10 + int((idx + 1) / len(candidate_ids) * 20)
            db.commit()

        run.current_agent = "Screener Agent"
        run.progress = 35
        db.commit()

        for idx, wd in enumerate(working_data):
            screened = agent_screen(wd["candidate"], wd["parsed"], job_description)
            wd["screened"] = screened
            run.screened_count = idx + 1
            run.progress = 35 + int((idx + 1) / len(working_data) * 25)
            db.commit()

        run.current_agent = "Deep Ranker Agent"
        run.progress = 65
        db.commit()

        working_data = agent_deep_rank(working_data, job_description)
        run.ranked_count = len(working_data)
        run.progress = 75
        db.commit()

        run.current_agent = "Finalizer Agent"
        run.progress = 85
        db.commit()

        working_data, best = agent_finalize(working_data, job_description)

        db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run.id).delete()
        db.commit()

        for idx, wd in enumerate(working_data):
            p = wd.get("parsed", {})
            s = wd.get("screened", {})
            r = wd.get("ranked", {})
            f = wd.get("final", {})
            skills_list = p.get("skills", [])
            skills_str = ", ".join(skills_list) if isinstance(skills_list, list) else str(skills_list)

            result = models.PipelineResult(
                run_id=run.id,
                candidate_id=wd["candidate_id"],
                candidate_name=p.get("name", wd["candidate"].get("name", "Unknown")),
                candidate_email=p.get("email", wd["candidate"].get("email", "")),
                role=p.get("role", wd["candidate"].get("role", "")),
                parsed_skills=skills_str,
                parsed_experience=p.get("experience_years"),
                parsed_location=p.get("location", ""),
                screened_score=s.get("screened_score"),
                screened_summary=s.get("screened_summary", ""),
                ranked_score=r.get("ranked_score"),
                ranked_analysis=r.get("ranked_analysis", ""),
                rank_position=r.get("rank_position", idx + 1),
                final_verdict=f.get("final_verdict", "Consider"),
                final_notes=f.get("final_notes", ""),
                is_best_match=(best is not None and wd["candidate_id"] == best["candidate_id"]),
            )
            db.add(result)

            orig = candidates_map.get(wd["candidate_id"])
            if orig:
                orig.match_score = r.get("ranked_score") or s.get("screened_score") or orig.match_score
                orig.current_stage = "Done"
                orig.summary = s.get("screened_summary", orig.summary)

        run.status = "completed"
        run.progress = 100
        run.current_agent = None
        run.completed_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(run)

    except Exception as e:
        run.status = "failed"
        run.error_message = str(e)[:500]
        db.commit()
        db.refresh(run)

    results = db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run.id).order_by(models.PipelineResult.rank_position).all()
    return schemas.PipelineFullOut(
        run=schemas.PipelineRunOut.model_validate(run),
        results=[schemas.PipelineResultOut.model_validate(r) for r in results],
    )


@router.get("/runs", response_model=list[schemas.PipelineRunOut])
def list_pipeline_runs(db: Session = Depends(get_db)):
    return db.query(models.PipelineRun).order_by(models.PipelineRun.created_at.desc()).all()


@router.get("/runs/{run_id}", response_model=schemas.PipelineFullOut)
def get_pipeline_run(run_id: str, db: Session = Depends(get_db)):
    run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id).first()
    if not run:
        raise HTTPException(404, "Pipeline run not found")
    results = db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run_id).order_by(models.PipelineResult.rank_position).all()
    return schemas.PipelineFullOut(
        run=schemas.PipelineRunOut.model_validate(run),
        results=[schemas.PipelineResultOut.model_validate(r) for r in results],
    )


@router.get("/runs/{run_id}/results")
def get_pipeline_results(run_id: str, db: Session = Depends(get_db)):
    results = db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run_id).order_by(models.PipelineResult.rank_position).all()
    return [schemas.PipelineResultOut.model_validate(r) for r in results]


@router.get("/results/{result_id}/export-txt")
def export_result_txt(result_id: str, db: Session = Depends(get_db)):
    r = db.query(models.PipelineResult).filter(models.PipelineResult.id == result_id).first()
    if not r:
        raise HTTPException(404, "Result not found")

    lines = [
        "=" * 60,
        f"AGENTIX AI - CANDIDATE EVALUATION REPORT",
        "=" * 60,
        "",
        f"Candidate: {r.candidate_name}",
        f"Email: {r.candidate_email or 'N/A'}",
        f"Role: {r.role or 'N/A'}",
        "",
        "-" * 40,
        "PARSED DATA",
        "-" * 40,
        f"Skills: {r.parsed_skills or 'N/A'}",
        f"Experience: {r.parsed_experience or 'N/A'} years",
        f"Location: {r.parsed_location or 'N/A'}",
        "",
        "-" * 40,
        "SCREENING RESULTS",
        "-" * 40,
        f"Score: {r.screened_score or 'N/A'}/100",
        f"Summary: {r.screened_summary or 'N/A'}",
        "",
        "-" * 40,
        "RANKING RESULTS",
        "-" * 40,
        f"Rank Position: #{r.rank_position or 'N/A'}",
        f"Ranked Score: {r.ranked_score or 'N/A'}/100",
        f"Analysis: {r.ranked_analysis or 'N/A'}",
        "",
        "-" * 40,
        "FINAL DECISION",
        "-" * 40,
        f"Verdict: {r.final_verdict or 'N/A'}",
        f"Notes: {r.final_notes or 'N/A'}",
        f"Best Match: {'Yes' if r.is_best_match else 'No'}",
        "",
        "=" * 60,
        "Generated by Agentix AI HR Manager",
        "=" * 60,
    ]
    content = "\n".join(lines)
    return StreamingResponse(
        io.StringIO(content),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.txt"},
    )


@router.get("/results/{result_id}/export-xlsx")
def export_result_xlsx(result_id: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    r = db.query(models.PipelineResult).filter(models.PipelineResult.id == result_id).first()
    if not r:
        raise HTTPException(404, "Result not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Report"

    ws.append(["AGENTIX AI - CANDIDATE EVALUATION REPORT"])
    ws.append([])
    ws.append(["Field", "Value"])
    ws.append(["Candidate", r.candidate_name])
    ws.append(["Email", r.candidate_email or "N/A"])
    ws.append(["Role", r.role or "N/A"])
    ws.append(["Skills", r.parsed_skills or "N/A"])
    ws.append(["Experience (years)", r.parsed_experience or "N/A"])
    ws.append(["Location", r.parsed_location or "N/A"])
    ws.append(["Screened Score", r.screened_score or "N/A"])
    ws.append(["Screening Summary", r.screened_summary or "N/A"])
    ws.append(["Rank Position", r.rank_position or "N/A"])
    ws.append(["Ranked Score", r.ranked_score or "N/A"])
    ws.append(["Rank Analysis", r.ranked_analysis or "N/A"])
    ws.append(["Final Verdict", r.final_verdict or "N/A"])
    ws.append(["Final Notes", r.final_notes or "N/A"])
    ws.append(["Best Match", "Yes" if r.is_best_match else "No"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.xlsx"},
    )


@router.get("/results/{result_id}/export-pdf")
def export_result_pdf(result_id: str, db: Session = Depends(get_db)):
    try:
        from fpdf import FPDF
    except ImportError:
        try:
            from fpdf2 import FPDF
        except ImportError:
            raise HTTPException(500, "fpdf2 not installed. Run: pip install fpdf2")

    r = db.query(models.PipelineResult).filter(models.PipelineResult.id == result_id).first()
    if not r:
        raise HTTPException(404, "Result not found")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Agentix AI - Candidate Evaluation Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Candidate: {r.candidate_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Email: {r.candidate_email or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Role: {r.role or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Parsed Data", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, f"Skills: {r.parsed_skills or 'N/A'}")
    pdf.cell(0, 5, f"Experience: {r.parsed_experience or 'N/A'} years", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Location: {r.parsed_location or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Screening Results", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Score: {r.screened_score or 'N/A'}/100", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Summary: {r.screened_summary or 'N/A'}")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Ranking Results", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Rank Position: #{r.rank_position or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Ranked Score: {r.ranked_score or 'N/A'}/100", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Analysis: {r.ranked_analysis or 'N/A'}")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Final Decision", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Verdict: {r.final_verdict or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Notes: {r.final_notes or 'N/A'}")
    pdf.cell(0, 5, f"Best Match: {'Yes' if r.is_best_match else 'No'}", new_x="LMARGIN", new_y="NEXT")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.pdf"},
    )
